from flask import Flask, request, jsonify, send_file, render_template, session
from flask_socketio import SocketIO, emit, join_room, leave_room, disconnect
from openpyxl import load_workbook
from flask_cors import CORS
from twilio.rest import Client
import imgkit
from jinja2 import Environment, FileSystemLoader
import os
import tempfile
import gspread
from gspread.utils import rowcol_to_a1
from oauth2client.service_account import ServiceAccountCredentials
from PIL import Image, ImageDraw, ImageFont, ImageOps
import qrcode
from datetime import datetime
import threading
import random

app = Flask(__name__)
CORS(app, origins='http://172.104.51.224')  # Adjust the port if necessary
# CORS(app, origins='http://139.59.127.215')
# CORS(app, resources={r"/*": {"origins": "*"}})
socketio = SocketIO(app, cors_allowed_origins="*")  # Allow CORS for all origins

app.secret_key = 'helloworld' 

# Twilio Configuration
TWILIO_ACCOUNT_SID = 'AC40da51365a0cf47a1fd992e5edceb727'  # Replace with your Twilio Account SID
TWILIO_AUTH_TOKEN = '6a8dcc1ad3fc6fd3cf6f0523e47c618d'    # Replace with your Twilio Auth Token
TWILIO_WHATSAPP_NUMBER = 'whatsapp:+6282348971093'  # Replace with your Twilio WhatsApp number

client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

# Google Sheets Configuration
SCOPE = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'service_account.json'  # Path to your service account JSON file

credentials = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, SCOPE)
gc = gspread.authorize(credentials)

# Open the Google Sheet
SHEET_NAME = 'DATA KURSI WORSHIP NIGHT'  # Replace with your Google Sheet name


try:
    sh = gc.open(SHEET_NAME)
    monitoring_ws = sh.worksheet('MONITORING 1')  # Replace with your worksheet name
    sendqueue_ws = sh.worksheet('SENDQUEUE')     # Open the SENDQUEUE worksheet
except Exception as e:
    print(f"Error accessing Google Sheet: {e}")
    exit(1)

@app.route('/occupied-seats', methods=['GET'])
def get_occupied_seats():
    """
    Retrieves a list of occupied seats from Google Sheets where only seats with 'SOLD' status are considered occupied.
    Also calculates and prints the quota units used.

    :return: JSON response with a list of occupied seat labels (e.g., ["A1", "B5", "C10"])
    """
    try:
        # Define the row and column where headers are located
        HEADER_ROW = 3  # Headers are in row 3
        START_COLUMN = 2  # Column B is 2

        # Fetch all values from the worksheet
        all_values = monitoring_ws.get_all_values()

        # Calculate the number of cells read
        num_rows = len(all_values)
        num_cols = max(len(row) for row in all_values) if all_values else 0
        total_cells_read = num_rows * num_cols

        # Calculate the quota units used
        quota_units_used = -(-total_cells_read // 100)  # Ceiling division

        # Print the quota usage information
        print(f"Total cells read: {total_cells_read}")
        print(f"Quota units used: {quota_units_used}")

        # [Rest of your code remains unchanged]

        # Check if the worksheet has enough rows
        if len(all_values) < HEADER_ROW:
            return jsonify({'occupied_seats': []}), 200

        # Extract headers from row 3, starting at column B
        header_row_values = all_values[HEADER_ROW - 1][START_COLUMN - 1:]

        # Ensure headers are unique
        if len(header_row_values) != len(set(header_row_values)):
            return jsonify({
                'success': False,
                'message': 'Header row contains duplicate headers. Please ensure all headers are unique.'
            }), 400

        # Create a header to column index mapping
        headers = {}
        for idx, header in enumerate(header_row_values, start=START_COLUMN):
            headers[header.strip().upper()] = idx  # idx corresponds to the column index (1-based)

        # Define required headers
        required_headers = ['BARIS', 'NO KURSI', 'SOLD/BOOK']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return jsonify({'success': False, 'message': f"Missing required headers: {', '.join(missing_headers)}"}), 400

        # Get column indices
        baris_col = headers['BARIS']
        seat_label_col = headers['NO KURSI']
        sold_book_status_col = headers['SOLD/BOOK']

        # Extract data starting from row 4
        data_rows = all_values[HEADER_ROW:]

        occupied_seats = []

        for row in data_rows:
            # Ensure the row has enough columns
            if len(row) < max(sold_book_status_col, seat_label_col, baris_col):
                continue

            sold_book_status = row[sold_book_status_col - 1].strip().upper()
            seat_number = row[seat_label_col - 1].strip()
            baris = row[baris_col - 1].strip().upper()

            # Check if baris and seat_number are not empty
            if not baris or not seat_number:
                continue

            seat_label = f"{baris}{seat_number}"

            # Only consider seats with 'SOLD' or 'BOOK' status as occupied
            if sold_book_status == 'SOLD' or sold_book_status == 'BOOK':
                occupied_seats.append(seat_label)


         # Now, get the list of currently locked seats
        current_locked_seats = list(locked_seats.keys())

        # Return both lists separately
        return jsonify({
            'occupied_seats': occupied_seats,
            'locked_seats': current_locked_seats
        }), 200

    except Exception as e:
        return jsonify({'success': False, 'message': f"Error retrieving occupied seats: {str(e)}"}), 500


# @app.route('/book', methods=['POST'])
# def book():
#     data = request.get_json()

#     if not data:
#         return jsonify({'success': False, 'message': 'No data provided.'}), 400

#     metadata = data.get('metadata', {})
#     bookings = data.get('bookings', [])
#     user_whatsapp = data.get('user_whatsapp', '')  # Ensure frontend sends user's WhatsApp number
#     print("user_whatsapp:", user_whatsapp)

#     # Validate user WhatsApp number
#     if not user_whatsapp.startswith('whatsapp:'):
#         return jsonify({'success': False, 'message': 'Invalid WhatsApp number format.'}), 400

#     # Extract metadata
#     movie_name = metadata.get('movieName', 'Unknown Movie')
#     show_date = metadata.get('showDate', 'N/A')
#     show_time = metadata.get('showTime', 'N/A')

#     if not bookings:
#         return jsonify({'success': False, 'message': 'No bookings provided.'}), 400

#     # Load the existing Excel file
#     # try:
#     #     workbook = load_workbook('DATA KURSI OPERA SAMADI.xlsx')
#     # except FileNotFoundError:
#     #     return jsonify({'success': False, 'message': 'Excel file not found.'}), 400

#     sheet = sh.worksheet('MONITORING 1')

#     # Get headers to find the correct columns
#     headers = {}
#     header_row = 3  # Headers are in the third row
#     first_data_row = header_row + 1  # Data starts from the row after headers

#     # Get the header row values
#     header_values = sheet.row_values(header_row)  # Fetch all values in the header row

#     # Read headers starting from column B (index 1 corresponds to column A)
#     for idx, cell_value in enumerate(header_values, start=1):  # start=1 to match Excel columns (A=1)
#         if idx >= 2:  # Skip column A (which is idx=1)
#             if cell_value:
#                 header_name = cell_value.strip().upper()
#                 headers[header_name] = idx  # idx corresponds to the Excel column number

#     # Debug: Print out the headers dictionary
#     print("Headers found in Excel:", headers)

#     required_headers = ['WARNA', 'BARIS', 'NO KURSI', 'NAMA', 'NO HP', 'SOLD/BOOK', 'QTY',
#                         'NILAI', 'TGL BAYAR', 'LUNAS', 'AMOUNT', 'PIC', 'GELANG', 'NO BANTU', 'KET', 'NO']

#     # Check if all required headers are present
#     missing_headers = [h for h in required_headers if h not in headers]
#     if missing_headers:
#         return jsonify({'success': False, 'message': f"Missing headers in Excel file: {', '.join(missing_headers)}"}), 400

#     # Find the next available NO
#     max_no = 0

#     # Assuming 'NO' is in Column A (1)
#     no_column_index = headers['NO']  # Should be 1 if 'NO' is in Column A

#     # Get all values in the 'NO' column starting from 'first_data_row'
#     no_column = worksheet.col_values(no_column_index)[first_data_row - 1:]  # gspread is 1-based, list is 0-based

#     # Iterate through the 'NO' values to find the maximum
#     for cell_no in no_column:
#         try:
#             cell_no_int = int(cell_no)
#             if cell_no_int > max_no:
#                 max_no = cell_no_int
#         except ValueError:
#             # If the cell is not an integer, skip it
#             continue

#     next_no = max_no + 1

#     # Collect booking details for the ticket
#     ticket_booking_details = {
#         'movie': movie_name,
#         'seats': [],
#         'date': show_date,
#         'time': show_time
#     }

#     # Iterate over each booking and update the Google Sheet
#     for booking in bookings:
#         warna = booking.get('WARNA', '').strip().upper()
#         baris = booking.get('BARIS', '').strip().upper()
#         no_kursi = str(booking.get('NO KURSI', '')).strip()

#         # Determine "Lunas" based on "SOLD/BOOK"
#         sold_book_status = booking.get('SOLD/BOOK', '').strip().upper()
#         if sold_book_status == 'SOLD':
#             lunas_status = 'Yes'
#         elif sold_book_status == 'BOOK':
#             lunas_status = 'No'
#         else:
#             return jsonify({'success': False, 'message': f"Invalid SOLD/BOOK status for seat {no_kursi}."}), 400

#         # Search for the row matching WARNA, BARIS, NO KURSI
#         found = False
#         no_kursi_cells = worksheet.findall(no_kursi, in_column=headers['NO KURSI'])
#         for cell in no_kursi_cells:
#             row = cell.row
#             cell_warna = worksheet.cell(row, headers['WARNA']).value
#             cell_baris = worksheet.cell(row, headers['BARIS']).value

#             if (str(cell_warna).strip().upper() == warna) and (str(cell_baris).strip().upper() == baris):
#                 # Found the correct row
#                 found = True

#                 # Update the row with booking details
#                 worksheet.update_cell(row, headers['NAMA'], booking.get('NAMA', ''))
#                 worksheet.update_cell(row, headers['NO HP'], booking.get('NO HP', ''))
#                 worksheet.update_cell(row, headers['SOLD/BOOK'], sold_book_status)
#                 worksheet.update_cell(row, headers['QTY'], booking.get('QTY', 1))
#                 worksheet.update_cell(row, headers['NILAI'], booking.get('NILAI', 0))
#                 worksheet.update_cell(row, headers['TGL BAYAR'], booking.get('Tgl Bayar', ''))
#                 worksheet.update_cell(row, headers['LUNAS'], lunas_status)
#                 worksheet.update_cell(row, headers['AMOUNT'], booking.get('Amount', 0))
#                 worksheet.update_cell(row, headers['PIC'], booking.get('PIC', ''))
#                 worksheet.update_cell(row, headers['GELANG'], booking.get('Gelang', ''))
#                 worksheet.update_cell(row, headers['NO BANTU'], booking.get('NO BANTU', ''))
#                 worksheet.update_cell(row, headers['KET'], booking.get('KET', ''))
#                 worksheet.update_cell(row, headers['NO'], next_no)  # Column A: NO

#                 # Increment NO for the next booking
#                 next_no += 1

#                 # Collect seat numbers for the ticket
#                 ticket_booking_details['seats'].append(no_kursi)

#                 break  # Exit the loop once the seat is found and updated

#         if not found:
#             return jsonify({'success': False, 'message': f"Seat {no_kursi} in row {baris} with color {warna} not found."}), 400

#     # Save the workbook
#     # try:
#     #     workbook.save('DATA KURSI OPERA SAMADI.xlsx')
#     # except Exception as e:
#     #     return jsonify({'success': False, 'message': f"Failed to save Excel file: {str(e)}"}), 500

#     # Generate the ticket image
#     try:
#         ticket_image_path = generate_ticket_image(ticket_booking_details)
#     except Exception as e:
#         return jsonify({'success': False, 'message': f"Failed to generate ticket image: {str(e)}"}), 500

#     # Send the ticket image via WhatsApp
#     # try:
#     #     # message = client.messages.create(
#     #     #     from_='whatsapp:+14155238886',  # Twilio Sandbox WhatsApp number
#     #     #     body='Hali ini TIKET anda dengan nomor 23623',
#     #     #     media_url=['http://45.77.39.229/tess.png'],
#     #     #     to='whatsapp:+6285320602318'  # Your verified phone number with country code
#     #     # )

#     #     # print(f"Message SID: {message.sid}")

#     #     # participant = client.conversations \
#     #     #             .v1 \
#     #     #             .conversations(conversation.sid) \
#     #     #             .participants \
#     #     #             .create(
#     #     #                 messaging_binding_address='whatsapp:+6285320602318',
#     #     #                 messaging_binding_proxy_address='whatsapp:+14155238886'
#     #     #             )

#     #     # print(participant.sid)

#     #     # message = client.conversations \
#     #     #         .v1 \
#     #     #         .conversations(conversation.sid) \
#     #     #         .messages \
#     #     #         .create(
#     #     #           author='system',
#     #     #           body='Here is your booking ticket!',
#     #     #           ContentSid='HXc500b4b9bbd245817efb6a925f86a0ff', 
#     #     #           ContentVariables='{"1":"A","2":"V"}'
#     #     #         )

#     #     # print(message.sid)
#     # except Exception as e:
#     #     return jsonify({'success': False, 'message': f"Failed to send WhatsApp message: {str(e)}"}), 500

#     # Optionally, delete the temporary image file after sending
#     # os.remove(ticket_image_path)

#     return jsonify({'success': True, 'message': 'Booking data saved and ticket sent via WhatsApp.'})

@app.route('/book', methods=['POST'])
def book():
    data = request.get_json()

    if not data:
        return jsonify({'success': False, 'message': 'No data provided.'}), 400

    username = data.get('username')  # Get the username from the request data

    if not username:
        return jsonify({'success': False, 'message': 'Username is required.'}), 401

    metadata = data.get('metadata', {})
    bookings = data.get('bookings', [])
    user_whatsapp = data.get('user_whatsapp', '')
    pic_whatsapp = data.get('pic_whatsapp', '')

    # Validate user WhatsApp number (optional)
    # if not user_whatsapp.startswith('whatsapp:'):
    #     return jsonify({'success': False, 'message': 'Invalid WhatsApp number format.'}), 400

    if not bookings:
        return jsonify({'success': False, 'message': 'No bookings provided.'}), 400

    # Extract metadata
    movie_name = metadata.get('movieName', 'Unknown Movie')
    show_date = metadata.get('showDate', 'N/A')
    show_time = metadata.get('showTime', 'N/A')

    # Fetch headers
    header_row = 3  # Headers are in the third row
    header_values = monitoring_ws.row_values(header_row)

    # Build headers dictionary (column name -> index)
    headers = {}
    for idx, cell_value in enumerate(header_values, start=1):
        if idx >= 2:  # Skip column A (idx=1)
            if cell_value:
                header_name = cell_value.strip().upper()
                headers[header_name] = idx  # 1-based index

    required_headers = ['WARNA', 'BARIS', 'NO KURSI', 'NAMA', 'NO HP', 'SOLD/BOOK', 'QTY',
                        'NILAI', 'TGL JUAL', 'TGL BAYAR', 'LUNAS', 'AMOUNT', 'PIC', 'GELANG', 'NO BANTU', 'KET', 'NO']

    # Check for missing headers
    missing_headers = [h for h in required_headers if h not in headers]
    if missing_headers:
        return jsonify({'success': False, 'message': f"Missing headers in Excel file: {', '.join(missing_headers)}"}), 400

    # Fetch all relevant data at once to minimize read operations
    all_records = monitoring_ws.get_all_records(head=header_row)

    # Build a mapping from (WARNA, BARIS, NO KURSI) to row number
    seat_to_row = {}
    for idx, record in enumerate(all_records, start=header_row + 1):
        key = (
            str(record.get('WARNA', '')).strip().upper(),
            str(record.get('BARIS', '')).strip().upper(),
            str(record.get('NO KURSI', '')).strip()
        )
        seat_to_row[key] = idx

    # Find the next available NO
    no_column_index = headers['NO']
    no_column = monitoring_ws.col_values(no_column_index)[header_row:]  # Skip header rows
    existing_nos = [int(no) for no in no_column if no.isdigit()]
    next_no = max(existing_nos, default=0) + 1

    # Prepare a list to hold all cell updates as dictionaries
    batch_update_requests = []

    # Collect booking details for the ticket
    first_booking = bookings[0]
    ticket_booking_details = {
        'NAMA': first_booking.get('NAMA', ''),
        'PIC': username,  # Use the username from the request data
        'seats': [],
        'date': show_date,
        'time': show_time
    }

    #ticket_id = generate_ticket_id(ticket_booking_details['seats'])  # Ensure this function is defined
    tickets = []
    for booking in bookings:
        ticket_id = generate_ticket_id(ticket_booking_details['seats']) 
        tickets.append(ticket_id)
        warna = booking.get('WARNA', '').strip().upper()
        baris = booking.get('BARIS', '').strip().upper()
        no_kursi = str(booking.get('NO KURSI', '')).strip()

        # Set booking status to 'SOLD' and Lunas to 'Yes'
        sold_book_status = 'SOLD'
        lunas_status = 'Yes'
        tgl_JUAL = datetime.now().strftime('%d/%m/%Y %H:%M:%S')  # Set current date

        key = (warna, baris, no_kursi)
        row = seat_to_row.get(key)

        if not row:
            return jsonify({'success': False, 'message': f"Seat {no_kursi} in row {baris} with color {warna} not found."}), 400

        # Prepare updates for this booking using A1 notation
        updates = {
            'NAMA': booking.get('NAMA', ''),
            'NO HP': booking.get('NO HP', ''),
            'SOLD/BOOK': sold_book_status,
            'QTY': booking.get('QTY', 1),
            'NILAI': booking.get('NILAI', 0),
            'TGL JUAL': tgl_JUAL,
            #'TGL BAYAR': tgl_bayar,
            #'LUNAS': lunas_status,
            # 'AMOUNT': booking.get('Amount', 0),
            'PIC': username,  # Use the username from the request data
            #'GELANG': booking.get('Gelang', ''),
            'KET': ticket_id,
            'NO': next_no
        }

        for field, value in updates.items():
            col_index = headers[field]
            col_letter = get_column_letter(col_index)
            cell_address = f"{col_letter}{row}"
            batch_update_requests.append({
                'range': cell_address,
                'values': [[value]]
            })

        # Increment NO for the next booking
        next_no += 1

        # Collect seat numbers for the ticket
        ticket_booking_details['seats'].append(baris + no_kursi)

    # Perform the batch update for MONITORING 1
    try:
        monitoring_ws.batch_update(batch_update_requests, value_input_option='USER_ENTERED')
    except Exception as e:
        return jsonify({'success': False, 'message': f"Failed to update MONITORING 1: {str(e)}"}), 500
    
    sendqueue_data = [
        #[ticket_id, user_whatsapp]
    ]

    # Generate the ticket image
    try:
        
        buyer_name = ticket_booking_details['NAMA']
        seller_name = "PIC  : " + ticket_booking_details['PIC']
        seat_label = ', '.join(ticket_booking_details['seats'])
        
        # Generate the ticket image
        for index, ticket_id in enumerate(tickets):
            sendqueue_data.append([ticket_id, user_whatsapp])
            sendqueue_data.append([ticket_id, pic_whatsapp])
            
            qr_data = f"{ticket_id}"
            generate_ticket_image(ticket_id, qr_data, buyer_name, seller_name, (ticket_booking_details['seats'][index]))
            seat_lss = ticket_booking_details['seats'][index]
            if seat_lss in locked_seats:
                del locked_seats[seat_lss]
    except Exception as e:
        return jsonify({'success': False, 'message': f"Failed to generate ticket image: {str(e)}"}), 500

    # Prepare data for SENDQUEUE
    # Assuming one ticket per request. If multiple, adjust accordingly.
    

    # Perform the bulk update for SENDQUEUE
    try:
        sendqueue_ws.append_rows(sendqueue_data, value_input_option='USER_ENTERED')
    except Exception as e:
        return jsonify({'success': False, 'message': f"Failed to update SENDQUEUE: {str(e)}"}), 500


    # Send the ticket image via WhatsApp
    # try:
    #     send_whatsapp_message(user_whatsapp, ticket_id)
    # except Exception as e:
    #     return jsonify({'success': False, 'message': f"Failed to send WhatsApp message: {str(e)}"}), 500

    # Optionally, delete the temporary image file after sending
    # os.remove(ticket_image_path)

    # Emit the 'seats_occupied' event to all connected clients
    try:
        socketio.emit('seats_occupied', {'seats': ticket_booking_details['seats']})
    except Exception as e:
        print(f"Error emitting 'seats_occupied' event: {e}")

    return jsonify({'success': True, 'message': 'Booking data saved and ticket generated.', 'id': ticket_id})

def get_column_letter(col_index):
    """
    Convert a 1-based column index to its corresponding Excel-style column letter.
    
    Parameters:
    - col_index (int): The 1-based column index (e.g., 1 for 'A', 2 for 'B', ..., 27 for 'AA').
    
    Returns:
    - str: The corresponding column letter(s).
    
    Example:
    >>> get_column_letter(1)
    'A'
    >>> get_column_letter(28)
    'AB'
    """
    if col_index < 1:
        raise ValueError("Column index must be 1 or greater.")

    letter = ''
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter

def send_whatsapp_message(to_whatsapp, image_path):
    # Implement your WhatsApp sending logic here, e.g., using Twilio API
    # Example using Twilio:

    # message = client.messages.create(
    #     body='Here is your booking ticket!',
    #     messaging_service_sid="MG42d8a53bbebd1b17d788dbe9686fe068",
    #     #media_url=[f'http://188.166.223.224/ticket/{image_path}.png'],
    #     to='+6285320602318'
    # )
    # print(f"{message}")
    
    pass

def wrap_text(text, font, max_width):
    lines = []
    words = text.split()
    line = ''
    for word in words:
        test_line = f"{line} {word}".strip()
        line_width, _ = font.getsize(test_line)
        if line_width <= max_width:
            line = test_line
        else:
            if line:
                lines.append(line)
            line = word
    if line:
        lines.append(line)
    return lines

def truncate_text(text, font, max_width):
    ellipsis = '...'
    ellipsis_width, _ = font.getsize(ellipsis)
    current_text = ''
    for char in text:
        test_text = current_text + char
        test_width, _ = font.getsize(test_text)
        if test_width + ellipsis_width <= max_width:
            current_text = test_text
        else:
            return current_text + ellipsis
    return current_text

def generate_ticket_id(seat_labels):
    """
    Generates a ticket ID by combining the seat labels and the current timestamp.

    :param seat_labels: List of seat labels (e.g., ["A11", "A12", "B2"])
    :return: The generated ticket ID as a string.
    """
    # Get the current date and time
    now = datetime.now()

    # Format the current time as YYYYMMDDHHMMSS
    timestamp = now.strftime('%Y%m%d%H%M%S')

    # Join the seat labels without any separators
    seats = ''.join(seat_labels)

    random_number = random.randint(1000, 9999)

    # Combine seats and timestamp to form the ticket ID
    ticket_id = f"{random_number}{timestamp}"

    return ticket_id

def classify_seat(seat):
    imamRows = ['A']
    vvipRows = ['B', 'C', 'D', 'E', 'F', 'G']
    specialRows = ['H', 'I', 'J', 'K']
    lastRow = ['L', 'M', 'N']
    moreRow = ['O', 'P', 'Q', 'R']

    # Extract row letter and seat number
    row_letter = ''.join(filter(str.isalpha, seat))
    seat_number = int(''.join(filter(str.isdigit, seat)))
    
    # Default seat class
    seat_class = 'blue'
    
    # Determine seat class based on row and seat number
    if row_letter in imamRows:
        if 8 <= seat_number <= 20:
            seat_class = 'red'
        else:
            seat_class = 'blue'
    elif row_letter in vvipRows:
        if 6 <= seat_number <= 25:
            seat_class = 'red'
        else:
            seat_class = 'blue'
    elif row_letter in specialRows:
        if 6 <= seat_number <= 25:
            seat_class = 'yellow'
        else:
            seat_class = 'blue'
    elif row_letter in lastRow:
        if 6 <= seat_number <= 25:
            seat_class = 'green'
        else:
            seat_class = 'blue'
    elif row_letter in moreRow:
        seat_class = 'blue'
    
    return seat_class

def generate_ticket_image(id, qr_data, buyer_name, seller_name, seat_label):
    """
    Generates a ticket image with specified details and returns the path to the saved image.

    :param id: Unique identifier used for naming the saved image file.
    :param qr_data: Text data to encode in the QR code.
    :param buyer_name: Name of the ticket buyer.
    :param seller_name: Name of the seller.
    :param seat_label: Seat label (e.g., "A1", "B4").
    :return: Path to the saved image file.
    """
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))

        seat_class = classify_seat(seat_label)
        fill_colors = seat_class

        if(seat_class == "blue"):
            fill_colors = (12,192,223)
        elif(seat_class == "green"):
            fill_colors = (0,214,28)
        elif(seat_class == "yellow"):
            fill_colors = (239,223,15)

        # Construct absolute paths
        image_path = os.path.join(script_dir, f'{seat_class}.png')
        font_path = os.path.join(script_dir, 'Poppins-Regular.ttf')
        
        
        # Open an existing image
        image = Image.open(image_path)
        
        # Convert image to RGB if necessary
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        # Create an ImageDraw object
        draw = ImageDraw.Draw(image)
        
        # Use a TrueType font
        font = ImageFont.truetype(font_path, size=45)
        
        # Verify image dimensions
        image_width, image_height = image.size
        
        colors = "white"
        
        # Define maximum text width
        max_text_width = 500  # Adjust this value based on your image
        
        # Starting positions
        x_position = 40
        y_positions = [1670, 1750, 1850]
        
        # Adjust the order of texts as per your request
        # New order: buyer_name, seller_name, seat_label
        texts = [buyer_name, ("SEAT : " + seat_label), seller_name]
        
        # Add text to the image
        for idx, text in enumerate(texts):
            y_position = y_positions[idx]
            if idx == 2:  # If it's the last text (seat_label), wrap text to new lines
                # lines = wrap_text(text, font, max_text_width)
                # for line in lines:
                #     draw.text((x_position, y_position), line, fill=colors, font=font)
                #     line_height = font.getsize(line)[1]
                #     y_position += line_height  # Move to next line
                truncated_text = truncate_text(text, font, max_text_width)
                draw.text((x_position, y_position), truncated_text, fill=colors, font=font)
            else:
                # For other texts, truncate with ellipsis if necessary
                truncated_text = truncate_text(text, ImageFont.truetype(font_path, size=50), max_text_width)
                draw.text((x_position, y_position), truncated_text, fill=colors, font=ImageFont.truetype(font_path, size=50))
        
        # Generate QR code
        qr = qrcode.QRCode(
            version=None,  # Controls the size of the QR code (1-40)
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,  # Controls how many pixels each "box" of the QR code is
            border=2,    # Controls how many boxes thick the border should be
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color=fill_colors).convert('RGB')
        
        # Resize QR code image if necessary
        qr_size = 380  # Desired size of the QR code
        qr_img = qr_img.resize((qr_size, qr_size), Image.ANTIALIAS)

        # qr_with_border = ImageOps.expand(qr_img, border=10, fill="black")
        
        # Position to paste the QR code on the main image
        qr_x = image_width - qr_size - 25  # Adjust 190 as needed for margin
        qr_y = image_height - qr_size - 20  # Adjust 140 as needed for margin
        
        # Paste the QR code onto the main image
        image.paste(qr_img, (qr_x, qr_y))

        # text_width, text_height = draw.textsize(seat_label, font=ImageFont.truetype(font_path, size=55))

        # text_position = (qr_x+(qr_img.width / 2)-(text_width/3), qr_y+(qr_img.height / 2)-(text_height/3))

        # draw.rectangle(
        #     [text_position[0]-8, text_position[1],
        #     text_position[0] + text_width, text_position[1] + text_height],
        #     fill=fill_colors
        # )
        # draw.text(text_position, seat_label, font=ImageFont.truetype(font_path, size=45), fill="black")

        # Define the output directory
        output_dir = 'ticket'

        # Create the directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        output_filename = os.path.join(output_dir, f'{id}.png')

        # Save the edited image
        image.save(output_filename)
        
        print("Text and QR code added to the image successfully.")
        
        return output_filename
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

@app.route('/show-ticket/<ticket_id>')
def show_ticket(ticket_id):
    # Assuming the ticket image is saved in the /static/ folder
    image_path = f"/ticket/{ticket_id}.png"
    
    # Pass the image path to the HTML template
    return render_template('ticket.html', image_path=image_path)

users = [
    {'username': 'aditya', 'password': '123123123', 'name' : 'Aditya Dharmaputra W', 'phone' : '6285320602318'},
    {'username': 'rivan', 'password': '123123123', 'name' : 'Rivan Tandiari', 'phone' : '6281355678777'},
    {'username': 'Evie', 'password': '12345678', 'name' : 'Evie Asvianty', 'phone' : '628194177993'},
    {'username': 'NellyHt', 'password': '5858', 'name' : 'Nelly', 'phone' : '6281355572999'},
    {'username': 'Lusyana', 'password': 'Jemmy88', 'name' : 'Lusyana', 'phone' : '6281343933800'},
    {'username': 'CR', 'password': '112233', 'name' : 'Erny Widjaja', 'phone' : '628114533476'},
    {'username': 'Rey', 'password': 'ymsby18', 'name' : 'Reynaldi S', 'phone' : '6282292931177'},
    {'username': 'yenni', 'password': '999999', 'name' : 'Yenni Ng', 'phone' : '628124271111'},
    {'username': 'Silvia', 'password': 'ABC123456', 'name' : 'Silvia Ira Wijaya', 'phone' : '62811460369'},
    {'username': 'lily', 'password': 'dgtompo15', 'name' : 'Lily', 'phone' : '628114182805'},
    {'username': 'Mey123', 'password': 'Happym3y', 'name' : 'Mey', 'phone' : '6289665640212'},
    {'username': 'richson', 'password': '888888', 'name' : 'Richson', 'phone' : '6281243224432'},
    {'username': 'melisa', 'password': '12345', 'name' : 'Melisa', 'phone' : '628111821185'},
]

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    # Check if the username and password match any user in the list
    user = next((user for user in users if user['username'] == username and user['password'] == password), None)
    
    if user:
        return jsonify({'success': True, 'username': username, 'name': user.get('name'), 'phone': user.get('phone')})
    else:
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

# Initialize a dictionary to keep track of locked seats
locked_seats = {}  # Key: seat_label, Value: {'user_id': socket_id, 'timestamp': datetime}

# Lock duration in seconds (e.g., 300 seconds = 5 minutes)
LOCK_DURATION = 300

# Lock expiration checker
def check_lock_expirations():
    threading.Timer(5, check_lock_expirations).start()  # Check every 5 seconds
    current_time = datetime.now()
    expired_seats = []
    for seat_label, lock_info in list(locked_seats.items()):
        lock_time = lock_info['timestamp']
        if (current_time - lock_time).total_seconds() > LOCK_DURATION:
            expired_seats.append(seat_label)
            del locked_seats[seat_label]
    if expired_seats:
        # Notify all clients about the unlocked seats
        socketio.emit('unlock_seats', {'seats': expired_seats})

# Start the lock expiration checker
check_lock_expirations()

# SocketIO events
@socketio.on('connect')
def handle_connect():
    print(f"Client connected: {request.sid}")
    # Send the list of currently locked and occupied seats
    current_locked_seats = list(locked_seats.keys())
    emit('initial_data', {
        'occupied_seats': get_occupied_seats_list(),
        'locked_seats': current_locked_seats
    })

def get_occupied_seats_list():
    # Function to retrieve occupied seats from Google Sheets
    # Similar to the logic in the '/occupied-seats' endpoint
    try:
        # Define the row and column where headers are located
        HEADER_ROW = 3  # Headers are in row 3
        START_COLUMN = 2  # Column B is 2

        # Fetch all values from the worksheet
        all_values = monitoring_ws.get_all_values()

        # Check if the worksheet has enough rows
        if len(all_values) < HEADER_ROW:
            return []

        # Extract headers from row 3, starting at column B
        header_row_values = all_values[HEADER_ROW - 1][START_COLUMN - 1:]

        # Create a header to column index mapping
        headers = {}
        for idx, header in enumerate(header_row_values, start=START_COLUMN):
            headers[header.strip().upper()] = idx

        # Get column indices
        baris_col = headers['BARIS']
        seat_label_col = headers['NO KURSI']
        sold_book_status_col = headers['SOLD/BOOK']

        # Extract data starting from row 4
        data_rows = all_values[HEADER_ROW:]

        occupied_seats = []

        for row in data_rows:
            if len(row) < max(sold_book_status_col, seat_label_col, baris_col):
                continue

            sold_book_status = row[sold_book_status_col - 1].strip().upper()
            seat_number = row[seat_label_col - 1].strip()
            baris = row[baris_col - 1].strip().upper()

            if not baris or not seat_number:
                continue

            seat_label = f"{baris}{seat_number}"

            if sold_book_status == 'SOLD':
                occupied_seats.append(seat_label)

        return occupied_seats

    except Exception as e:
        print(f"Error retrieving occupied seats: {str(e)}")
        return []

@socketio.on('disconnect')
def handle_disconnect():
    print(f"Client disconnected: {request.sid}")
    # Release any seats locked by this user
    released_seats = []
    for seat_label, lock_info in list(locked_seats.items()):
        if lock_info['user_id'] == request.sid:
            del locked_seats[seat_label]
            released_seats.append(seat_label)
    if released_seats:
        # Notify all clients about the unlocked seats
        socketio.emit('unlock_seats', {'seats': released_seats})

@socketio.on('lock_seat')
def handle_lock_seat(data):
    seat_label = data['seat']
    if seat_label in locked_seats:
        # Seat is already locked
        emit('seat_lock_failed', {'seat': seat_label})
    else:
        # Lock the seat
        locked_seats[seat_label] = {'user_id': request.sid, 'timestamp': datetime.now()}
        # Notify all clients about the locked seat
        socketio.emit('lock_seats', {'seats': [seat_label]}, include_self=False)

@socketio.on('unlock_seat')
def handle_unlock_seat(data):
    seat_label = data['seat']
    if seat_label in locked_seats and locked_seats[seat_label]['user_id'] == request.sid:
        # Unlock the seat
        del locked_seats[seat_label]
        # Notify all clients about the unlocked seat
        socketio.emit('unlock_seats', {'seats': [seat_label]}, include_self=False)

@app.route('/verify-transaction', methods=['POST'])
def verify_transaction():
    data = request.get_json()
    transaction_id = data.get('transaction_id', '').strip()

    if not transaction_id:
        return jsonify({'success': False, 'message': 'No transaction ID provided.'}), 400

    try:
        # Fetch the headers from the sheet (read operation)
        header_row = 3  # Headers are in the third row
        header_values = monitoring_ws.row_values(header_row)

        # Build headers dictionary (column name -> index)
        headers = {}
        for idx, cell_value in enumerate(header_values, start=1):
            if cell_value:
                header_name = cell_value.strip().upper()
                headers[header_name] = idx  # 1-based index

        if 'KET' not in headers:
            return jsonify({'success': False, 'message': 'KET column not found in the sheet.'}), 500

        # Find all cells with the transaction_id in the 'KET' column
        ket_col_index = headers['KET']  # 1-based index

        try:
            matching_cells = monitoring_ws.findall(transaction_id, in_column=ket_col_index)
        except Exception as e:
            print(f"Error finding transaction ID: {e}")
            matching_cells = []

        if matching_cells:
            transaction_details = []
            for cell in matching_cells:
                # Fetch the entire row for each matching cell (read operation per row)
                row_values = monitoring_ws.row_values(cell.row)

                # Define the fields you want to include in the response
                fields_to_include = ['WARNA', 'BARIS', 'NO KURSI', 'NAMA', 'NO HP', 
                                     'PIC', 'GELANG']

                # Build a dictionary containing the transaction details
                transaction = {}
                for header in fields_to_include:
                    col_index = headers.get(header)
                    if col_index:
                        col_idx = col_index - 1  # Zero-based index
                        if col_idx < len(row_values):
                            transaction[header] = row_values[col_idx]
                        else:
                            transaction[header] = ''
                    else:
                        transaction[header] = ''

                transaction_details.append(transaction)

            return jsonify({
                'success': True, 
                'message': f'Found {len(transaction_details)} transaction(s) with ID {transaction_id}.', 
                'details': transaction_details
            })
        else:
            return jsonify({'success': False, 'message': 'Transaction ID not found.'}), 404

    except Exception as e:
        print(f"Error verifying transaction: {e}")
        return jsonify({'success': False, 'message': 'An error occurred during verification.'}), 500

def get_headers():
    header_row = 3
    header_values = monitoring_ws.row_values(header_row)
    headers = {}
    for idx, cell_value in enumerate(header_values, start=1):
        if cell_value:
            header_name = cell_value.strip().upper()
            headers[header_name] = idx
    return headers

@app.route('/give-gelang', methods=['POST'])
def give_gelang():
    data = request.get_json()
    transaction_id = data.get('transaction_id', '').strip()

    if not transaction_id:
        return jsonify({'success': False, 'message': 'No transaction ID provided.'}), 400

    try:
        headers = get_headers()

        if 'KET' not in headers or 'GELANG' not in headers:
            return jsonify({'success': False, 'message': 'Required columns not found in the sheet.'}), 500

        ket_col_index = headers['KET']
        gelang_col_index = headers['GELANG']

        # Find all cells with the transaction_id in the 'KET' column
        matching_cells = monitoring_ws.findall(transaction_id, in_column=ket_col_index)
        if not matching_cells:
            return jsonify({'success': False, 'message': 'Transaction ID not found.'}), 404

        # Collect all row numbers that match the transaction_id
        matching_row_numbers = [cell.row for cell in matching_cells]

        # Prepare A1 notation ranges for batch_get (to check current GELANG status)
        ranges = [f"{rowcol_to_a1(row_num, gelang_col_index)}:{rowcol_to_a1(row_num, gelang_col_index)}" for row_num in matching_row_numbers]

        # Use batch_get to fetch current GELANG statuses
        try:
            batch_gelang_values = monitoring_ws.batch_get(ranges)
        except Exception as e:
            print(f"Error fetching GELANG statuses via batch_get: {e}")
            return jsonify({'success': False, 'message': 'Failed to fetch current GELANG statuses.'}), 500

        cells_to_update = []
        for row_num, gelang_values in zip(matching_row_numbers, batch_gelang_values):
            # gelang_values is a list of lists (rows), each containing cell values
            current_gelang = gelang_values[0][0] if gelang_values and gelang_values[0] else ''
            if current_gelang.strip().upper() != 'YES':
                cell_address = rowcol_to_a1(row_num, gelang_col_index)
                cells_to_update.append({
                    'range': cell_address,
                    'values': [['Yes']]
                })

        if not cells_to_update:
            return jsonify({'success': False, 'message': 'GELANG is already marked as Yes for all matching transactions.'}), 400

        # Perform batch update to set GELANG to 'Yes'
        try:
            monitoring_ws.batch_update(cells_to_update)
        except Exception as e:
            print(f"Error updating GELANG statuses: {e}")
            return jsonify({'success': False, 'message': 'Failed to update GELANG statuses.'}), 500

        return jsonify({
            'success': True,
            'message': f'GELANG has been marked as Yes for {len(cells_to_update)} transaction(s).'
        }), 200

    except Exception as e:
        print(f"Error giving Gelang: {e}")
        return jsonify({'success': False, 'message': 'An error occurred while updating GELANG.'}), 500

@app.route('/verify')
def verify():
    return render_template('verify.html')

if __name__ == "__main__":
    socketio.run(app, host='0.0.0.0', port=5000)
    # socketio.run(app, host='0.0.0.0', port=5000, ssl_context=('/etc/ssl/private/selfsigned.crt', '/etc/ssl/private/selfsigned.key'))
